"""
Generates data.xlsx for AI Board.
Run from the AI_Board folder:  python create_excel.py

Sheet structure (matches parseWorkbook in data.jsx):
  Initiatives    — one row per initiative
  Milestones     — one row per milestone (linked by initiative id)
  Business_Units — BU catalogue
  Technologies   — technology catalogue
  Blockers       — blocker catalogue
  Platforms      — platform catalogue (BU-owned grouping layer)
  Help           — usage guide
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

# ── Catalogue data ─────────────────────────────────────────────────────────────

BUSINESS_UNITS = [
    # id,    name,           short, accent (OKLCH),              lead
    ("mkt", "Marketing",      "MK", "oklch(0.48 0.12 290)", "Anne Berg"),
    ("cxo", "Cust. Ops",      "CO", "oklch(0.48 0.12 165)", "Mads Iversen"),
    ("dig", "Digital & Data", "DD", "oklch(0.55 0.13 55)",  "Sofie Lund"),
]

PLATFORMS = [
    # id,          name,       business unit id
    ("p_adobe",   "Adobe",    "mkt"),
    ("p_genesys", "Genesys",  "cxo"),
    ("p_boost",   "Boost.ai", "cxo"),
]

TECHNOLOGIES = [
    # id,           name,               category,        color hue
    ("t_gpt",     "GPT-4o",           "LLM",           250),
    ("t_azure",   "Azure OpenAI",     "LLM",           220),
    ("t_rag",     "RAG / Vector DB",  "LLM",           290),
    ("t_whisper", "Whisper STT",      "Speech",         30),
    ("t_xcript",  "AutoTranscript",   "Speech",         50),
    ("t_aa",      "Agent Assist",     "Agent tooling", 180),
    ("t_auto",    "Agent Automation", "Agent tooling", 200),
    ("t_copilot", "Copilot M365",     "Assistant",     150),
]

BLOCKERS = [
    # id,          name,                   category,         color hue
    ("b_legal",  "Legal review",          "Godkendelse",    10),
    ("b_budget", "Budget godkendelse",    "Godkendelse",    30),
    ("b_gdpr",   "GDPR compliance",       "Godkendelse",     5),
    ("b_it",     "IT infrastruktur",      "Teknik",        200),
    ("b_vendor", "Vendor kontrakt",       "Teknik",        220),
    ("b_data",   "Data adgang",           "Teknik",        240),
    ("b_talent", "Kompetencer",           "Organisation",   60),
    ("b_res",    "Ressourcer/kapacitet",  "Organisation",   70),
]

# Columns: id, business unit id, platform id, name, status, owner, start date, end date,
#          description, tags, technologies, blockers
INITIATIVES = [
    # fmt: off
    ("i_camp",    "mkt", "p_adobe",   "Personalised campaign generation",   "poc",   "Marketing Ops",   "2026-03-01", "2026-09-30",  "Generate variant copy + creative per segment using LLMs.", "B2C, EU",         "GPT-4o, Azure OpenAI",                       "Legal review, GDPR compliance"),
    ("i_tag",     "mkt", "p_adobe",   "Asset auto-tagging",                 "live",  "DAM team",        "2025-11-01", "2026-12-31",  "Classify and tag the DAM stock library automatically.",    "Content ops",     "GPT-4o, RAG / Vector DB",                    ""),
    ("i_seo",     "mkt", "",          "SEO content assistant",              "idea",  "Brand studio",    "2026-09-01", "2027-02-28",  "Draft SEO-optimised articles from briefs and keywords.",   "Content, SEO",    "GPT-4o, Azure OpenAI",                       "Budget godkendelse"),
    ("i_persona", "mkt", "",          "Customer persona engine",            "pilot", "Insights",        "2026-04-01", "2026-11-30",  "Build dynamic customer segments from behavioural data.",   "Analytics, B2C",  "RAG / Vector DB",                            "Kompetencer"),
    ("i_voice",   "cxo", "p_boost",   "Customer voicebot",                  "pilot", "Customer Care",   "2026-02-01", "2026-11-30",  "Voice-first front door for the service line.",             "B2C, 24/7",       "Whisper STT, Agent Automation, Agent Assist", "Legal review, IT infrastruktur"),
    ("i_helpdesk","cxo", "p_genesys", "IT helpdesk bot",                    "live",  "IT",              "2025-09-01", "2026-12-31",  "Tier-0 IT support via Slack and employee portal.",         "Internal",        "GPT-4o, Agent Assist, Agent Automation",      "GDPR compliance"),
    ("i_sum",     "cxo", "",          "Call summarisation",                 "live",  "Contact Centre",  "2025-12-01", "2026-12-31",  "Auto-summarise calls and write them into CRM.",            "Productivity",    "Whisper STT, AutoTranscript",                ""),
    ("i_qa",      "cxo", "p_genesys", "Agent-assist QA",                    "pilot", "Quality",         "2026-03-01", "2026-10-31",  "Score 100% of interactions against quality rubrics.",      "Compliance",      "AutoTranscript, Agent Assist",               "Vendor kontrakt"),
    ("i_search",  "dig", "",          "Enterprise knowledge search",        "poc",   "Platform team",   "2026-04-01", "2026-12-31",  "RAG-based search across internal wikis and documents.",    "Internal",        "GPT-4o, RAG / Vector DB",                    "Legal review, Vendor kontrakt, Data adgang"),
    ("i_trans",   "dig", "",          "Meeting transcription",              "live",  "Workplace IT",    "2026-01-01", "2026-12-31",  "Transcribe, summarise, and action-item all meetings.",     "Productivity",    "Whisper STT, AutoTranscript, Copilot M365",  "IT infrastruktur"),
    ("i_assist",  "dig", "",          "Developer copilot",                  "pilot", "Engineering",     "2026-03-01", "2026-09-30",  "In-IDE code suggestions tuned to internal patterns.",      "DevEx",           "GPT-4o, Copilot M365, Azure OpenAI",         "Kompetencer"),
    ("i_predict", "dig", "",          "Predictive analytics engine",        "idea",  "Data Science",    "2026-09-01", "2027-03-31",  "Forecast churn and NPS from interaction and CRM data.",   "Analytics",       "RAG / Vector DB, Agent Automation",          "Budget godkendelse, Data adgang, Ressourcer/kapacitet"),
    # fmt: on
]

# Columns: initiative id, date, label
MILESTONES = [
    ("i_camp",    "2026-06-15", "Pilot kickoff"),
    ("i_tag",     "2026-04-10", "Go-live"),
    ("i_persona", "2026-07-01", "Pilot live"),
    ("i_voice",   "2026-07-01", "Pilot live"),
    ("i_helpdesk","2026-02-01", "Go-live"),
    ("i_helpdesk","2026-08-15", "EN+DK rollout"),
    ("i_sum",     "2026-03-20", "Go-live"),
    ("i_qa",      "2026-06-01", "Pilot live"),
    ("i_search",  "2026-10-15", "POC review"),
    ("i_trans",   "2026-03-01", "Go-live"),
    ("i_assist",  "2026-06-01", "Pilot live"),
]

# ── Style constants ────────────────────────────────────────────────────────────

STATUS_FILLS = {
    "idea":  PatternFill("solid", fgColor="F2F2F0"),
    "poc":   PatternFill("solid", fgColor="EBF0FB"),
    "pilot": PatternFill("solid", fgColor="FDF5E6"),
    "live":  PatternFill("solid", fgColor="EDFAF3"),
}
HDR_FILL  = PatternFill("solid", fgColor="1A1A1A")
HDR_FONT  = Font(color="FFFFFF", bold=True, size=10)
MONO_FONT = Font(name="Consolas", size=9, color="555555")
BORDER    = Border(
    left=Side(style="thin", color="E0E0DC"), right=Side(style="thin", color="E0E0DC"),
    top=Side(style="thin", color="E0E0DC"),  bottom=Side(style="thin", color="E0E0DC"),
)

# ── Helpers ────────────────────────────────────────────────────────────────────

def write_header(ws, headers_with_comments):
    ws.row_dimensions[1].height = 24
    for col, (header, comment_text) in enumerate(headers_with_comments, 1):
        c = ws.cell(row=1, column=col, value=header)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(vertical="center", horizontal="left")
        c.border = BORDER
        if comment_text:
            cmt = Comment(comment_text, "AI Board"); cmt.width, cmt.height = 260, 100
            c.comment = cmt

def style_data_row(ws, row_idx, ncols, fill=None):
    for col in range(1, ncols + 1):
        c = ws.cell(row=row_idx, column=col)
        c.border = BORDER
        c.alignment = Alignment(vertical="center", horizontal="left", wrap_text=False)
        if fill: c.fill = fill

def auto_width(ws, headers, rows, max_w=55):
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(header)
        for row in rows:
            val = str(row[col_idx - 1]) if col_idx <= len(row) and row[col_idx - 1] is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 3, max_w)

# ── Sheet: Initiatives ────────────────────────────────────────────────────────

def make_initiatives_sheet(wb):
    ws = wb.create_sheet("Initiatives")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    bu_ids    = [r[0] for r in BUSINESS_UNITS]
    plat_ids  = [r[0] for r in PLATFORMS]

    HEADERS = [
        ("ID",              "Unique key (e.g. i_001). No spaces or special chars."),
        ("Business Unit ID","ID from Business_Units sheet. Valid: " + ", ".join(bu_ids)),
        ("Platform ID",     "Optional ID from Platforms sheet. Leave blank if not on a platform. Valid: " + ", ".join(plat_ids)),
        ("Name",            "Short descriptive title."),
        ("Status",          "idea / poc / pilot / live"),
        ("Owner",           "Name or team responsible."),
        ("Start Date",      "Format: YYYY-MM-DD"),
        ("End Date",        "Format: YYYY-MM-DD"),
        ("Description",     "Brief description of purpose and scope."),
        ("Tags",            "Comma-separated keywords, e.g.: B2C, EU, Internal"),
        ("Technologies",    "Comma-separated names from Technologies sheet."),
        ("Blockers",        "Comma-separated names from Blockers sheet. Leave blank if none."),
    ]
    write_header(ws, HEADERS)
    headers_plain = [h for h, _ in HEADERS]

    dv_status = DataValidation(type="list", formula1='"idea,poc,pilot,live"', allow_blank=True,
        showErrorMessage=True, errorTitle="Invalid status", error="Choose: idea, poc, pilot or live")
    dv_bu = DataValidation(type="list", formula1='"' + ",".join(bu_ids) + '"', allow_blank=True,
        showErrorMessage=True, errorTitle="Invalid BU", error="Choose a valid BU ID.")
    ws.add_data_validation(dv_status); ws.add_data_validation(dv_bu)

    for row_idx, row in enumerate(INITIATIVES, 2):
        fill = STATUS_FILLS.get(row[4].lower())
        values = list(row)
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.border = BORDER
            c.alignment = Alignment(vertical="center", horizontal="left", wrap_text=False)
            if fill: c.fill = fill
            if col == 1: c.font = MONO_FONT
        ws.row_dimensions[row_idx].height = 19
        dv_status.add(ws.cell(row=row_idx, column=5))
        dv_bu.add(ws.cell(row=row_idx, column=2))

    # Empty rows for new entries
    last_row = len(INITIATIVES) + 2
    for row_idx in range(last_row, last_row + 10):
        for col in range(1, len(HEADERS) + 1):
            ws.cell(row=row_idx, column=col).border = BORDER
        ws.row_dimensions[row_idx].height = 19
        dv_status.add(ws.cell(row=row_idx, column=5))
        dv_bu.add(ws.cell(row=row_idx, column=2))

    auto_width(ws, headers_plain, [list(r) for r in INITIATIVES])
    ws.column_dimensions["I"].width = 50  # Description
    ws.column_dimensions["K"].width = 35  # Technologies
    ws.column_dimensions["L"].width = 35  # Blockers
    return ws

# ── Sheet: Milestones ─────────────────────────────────────────────────────────

def make_milestones_sheet(wb):
    ws = wb.create_sheet("Milestones")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    HEADERS = [
        ("Initiative ID", "ID of the parent initiative from the Initiatives sheet."),
        ("Date",          "Format: YYYY-MM-DD"),
        ("Label",         "Short milestone name, e.g.: Go-live, Pilot kickoff"),
    ]
    write_header(ws, HEADERS)

    for row_idx, row in enumerate(MILESTONES, 2):
        for col, val in enumerate(row, 1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.border = BORDER
            c.alignment = Alignment(vertical="center", horizontal="left")
            if col == 1: c.font = MONO_FONT
        ws.row_dimensions[row_idx].height = 19

    auto_width(ws, [h for h, _ in HEADERS], [list(r) for r in MILESTONES])
    return ws

# ── Sheet: Business_Units ─────────────────────────────────────────────────────

def make_bu_sheet(wb):
    ws = wb.create_sheet("Business_Units")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    HEADERS = [
        ("ID",           "Unique key used in Initiatives sheet."),
        ("Name",         "Full name of the business unit."),
        ("Short",        "2-3 letter abbreviation shown on the board."),
        ("Accent Color", "OKLCH color string used on the board. Leave as-is."),
        ("Lead",         "Name of the responsible leader."),
    ]
    write_header(ws, HEADERS)

    for row_idx, row in enumerate(BUSINESS_UNITS, 2):
        for col, val in enumerate(row, 1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.border = BORDER
            c.alignment = Alignment(vertical="center", horizontal="left")
            if col == 1: c.font = MONO_FONT
        ws.row_dimensions[row_idx].height = 19

    auto_width(ws, [h for h, _ in HEADERS], [list(r) for r in BUSINESS_UNITS])
    return ws

# ── Sheet: Platforms ──────────────────────────────────────────────────────────

def make_platforms_sheet(wb):
    ws = wb.create_sheet("Platforms")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    bu_ids = [r[0] for r in BUSINESS_UNITS]

    HEADERS = [
        ("ID",              "Unique key used in Initiatives sheet under 'Platform ID'."),
        ("Name",            "Platform name shown as a sub-header on the board, e.g.: Adobe, Genesys, Boost.ai"),
        ("Business Unit ID","ID from Business_Units sheet. Valid: " + ", ".join(bu_ids)),
    ]
    write_header(ws, HEADERS)

    dv_bu = DataValidation(type="list", formula1='"' + ",".join(bu_ids) + '"', allow_blank=True)
    ws.add_data_validation(dv_bu)

    for row_idx, row in enumerate(PLATFORMS, 2):
        for col, val in enumerate(row, 1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.border = BORDER
            c.alignment = Alignment(vertical="center", horizontal="left")
            if col == 1: c.font = MONO_FONT
        ws.row_dimensions[row_idx].height = 19
        dv_bu.add(ws.cell(row=row_idx, column=3))

    # Empty rows for new entries
    last_row = len(PLATFORMS) + 2
    for row_idx in range(last_row, last_row + 5):
        for col in range(1, len(HEADERS) + 1):
            ws.cell(row=row_idx, column=col).border = BORDER
        ws.row_dimensions[row_idx].height = 19
        dv_bu.add(ws.cell(row=row_idx, column=3))

    auto_width(ws, [h for h, _ in HEADERS], [list(r) for r in PLATFORMS])
    return ws

# ── Sheet: Technologies ───────────────────────────────────────────────────────

def make_tech_sheet(wb):
    ws = wb.create_sheet("Technologies")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    HEADERS = [
        ("ID",         "Internal key. Do not change."),
        ("Name",       "Technology name. Used in Initiatives sheet under 'Technologies'."),
        ("Category",   "Grouping category on the board, e.g. LLM, Speech, Agent tooling."),
        ("Color Hue",  "Optional: number 0-360 for color on the board. Can be blank."),
    ]
    write_header(ws, HEADERS)

    for row_idx, row in enumerate(TECHNOLOGIES, 2):
        for col, val in enumerate(row, 1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.border = BORDER
            c.alignment = Alignment(vertical="center", horizontal="left")
            if col == 1: c.font = MONO_FONT
        ws.row_dimensions[row_idx].height = 19

    auto_width(ws, [h for h, _ in HEADERS], [list(r) for r in TECHNOLOGIES])
    return ws

# ── Sheet: Blockers ───────────────────────────────────────────────────────────

def make_blockers_sheet(wb):
    ws = wb.create_sheet("Blockers")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    HEADERS = [
        ("ID",         "Internal key. Do not change."),
        ("Name",       "Blocker name. Used in Initiatives sheet under 'Blockers'."),
        ("Category",   "Grouping category, e.g. Godkendelse, Teknik, Organisation."),
        ("Color Hue",  "Optional: number 0-360. Can be blank."),
    ]
    write_header(ws, HEADERS)

    for row_idx, row in enumerate(BLOCKERS, 2):
        for col, val in enumerate(row, 1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.border = BORDER
            c.alignment = Alignment(vertical="center", horizontal="left")
            if col == 1: c.font = MONO_FONT
        ws.row_dimensions[row_idx].height = 19

    auto_width(ws, [h for h, _ in HEADERS], [list(r) for r in BLOCKERS])
    return ws

# ── Sheet: Help ───────────────────────────────────────────────────────────────

def make_help_sheet(wb):
    ws = wb.create_sheet("Help")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 70

    def hdr(r, txt):
        c = ws.cell(row=r, column=1, value=txt)
        c.font = Font(color="FFFFFF", bold=True, size=11)
        c.fill = HDR_FILL
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        ws.row_dimensions[r].height = 20
        return r + 1

    def row(r, a, b="", bold_a=False, fill=None):
        ca = ws.cell(row=r, column=1, value=a)
        cb = ws.cell(row=r, column=2, value=b)
        if bold_a: ca.font = Font(bold=True, size=11)
        if fill: ca.fill = fill; cb.fill = fill
        ca.alignment = Alignment(vertical="top", wrap_text=True)
        cb.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = max(15, len(str(b)) // 3)
        return r + 1

    r = 1
    r = hdr(r, "AI BOARD — DATA GUIDE")
    r = row(r, "")
    r = hdr(r, "HOW TO ADD AN INITIATIVE")
    r = row(r, "1. Sheet",      "Go to the 'Initiatives' sheet")
    r = row(r, "2. Row",        "Find the next empty row below existing initiatives")
    r = row(r, "3. ID",         "Give it a unique ID, e.g.: i_my_initiative")
    r = row(r, "4. BU",         "Choose a Business Unit ID from the dropdown")
    r = row(r, "5. Platform",   "Optional: choose a Platform ID if it belongs to one")
    r = row(r, "6. Status",     "Choose from dropdown: idea / poc / pilot / live")
    r = row(r, "7. Dates",      "Format: YYYY-MM-DD, e.g. 2026-06-01")
    r = row(r, "8. Tech",       "Comma-separated names from the Technologies sheet")
    r = row(r, "9. Blockers",   "Comma-separated names from Blockers sheet (or leave blank)")
    r = row(r, "10. Save",      "Save the file, go to the board and click Import")
    r = row(r, "")
    r = hdr(r, "HOW TO ADD A MILESTONE")
    r = row(r, "Sheet",         "Go to the 'Milestones' sheet")
    r = row(r, "Initiative ID", "Enter the ID of the parent initiative")
    r = row(r, "Date",          "Format: YYYY-MM-DD")
    r = row(r, "Label",         "Short name, e.g.: Go-live, Pilot kickoff")
    r = row(r, "")
    r = hdr(r, "HOW TO ADD A PLATFORM")
    r = row(r, "Sheet",         "Go to the 'Platforms' sheet")
    r = row(r, "ID",            "Unique key, e.g.: p_salesforce")
    r = row(r, "Name",          "Display name, e.g.: Salesforce")
    r = row(r, "Business Unit ID", "ID of the owning BU from Business_Units sheet")
    r = row(r, "")
    r = hdr(r, "STATUS COLOURS")
    r = row(r, "idea",  "Idea phase — not started yet",         fill=STATUS_FILLS["idea"])
    r = row(r, "poc",   "Proof of Concept — early testing",     fill=STATUS_FILLS["poc"])
    r = row(r, "pilot", "Pilot phase — limited rollout",        fill=STATUS_FILLS["pilot"])
    r = row(r, "live",  "In production",                        fill=STATUS_FILLS["live"])
    r = row(r, "")
    r = hdr(r, "VALID BUSINESS UNIT IDs")
    for bu in BUSINESS_UNITS:
        r = row(r, bu[0], f"{bu[1]}  (Lead: {bu[4]})")
    r = row(r, "")
    r = hdr(r, "VALID PLATFORM IDs")
    for p in PLATFORMS:
        r = row(r, p[0], f"{p[1]}  (BU: {p[2]})")
    r = row(r, "")
    r = hdr(r, "VALID TECHNOLOGY NAMES")
    for t in TECHNOLOGIES:
        r = row(r, t[2], t[1])
    r = row(r, "")
    r = hdr(r, "VALID BLOCKER NAMES")
    for b in BLOCKERS:
        r = row(r, b[2], b[1])

    return ws

# ── Build workbook ─────────────────────────────────────────────────────────────

wb = openpyxl.Workbook()
wb.remove(wb.active)

make_initiatives_sheet(wb)
make_milestones_sheet(wb)
make_bu_sheet(wb)
make_platforms_sheet(wb)
make_tech_sheet(wb)
make_blockers_sheet(wb)
make_help_sheet(wb)

wb.active = wb["Initiatives"]
wb.save("data.xlsx")

print("OK  data.xlsx written.")
print(f"   {len(BUSINESS_UNITS)} business units, {len(PLATFORMS)} platforms,")
print(f"   {len(TECHNOLOGIES)} technologies, {len(BLOCKERS)} blockers,")
print(f"   {len(INITIATIVES)} initiatives, {len(MILESTONES)} milestones.")
print()
print("Sheets: Initiatives, Milestones, Business_Units, Platforms, Technologies, Blockers, Help")
